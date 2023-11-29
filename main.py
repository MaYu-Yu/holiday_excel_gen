import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, colors, Border, Side

def generate_excel(csv_path, out_path, title, header, column_widths):
    # 讀取CSV檔案
    data = []
    if not os.path.exists(out_path):
        os.makedirs(out_path)
    out_path = os.path.join(out_path, f'{title}.xlsx')

    # 開始處理
    with open(csv_path, 'r', encoding='big5') as file:
        reader = csv.DictReader(file)
        for row in reader:
            data.append(row)

    # 建立Excel工作簿
    workbook = Workbook()

    # 解析CSV資料並生成Excel檔案
    for month in range(1, 13):
        # 過濾當月的資料
        month_data = [row for row in data if int(row['西元日期'][4:6]) == month]
        if not month_data:
            continue

        # 創建工作表
        sheet_name = f'{month}月'
        sheet = workbook.create_sheet(title=sheet_name)

        # 設定列寬
        for column, width in column_widths.items():
            sheet.column_dimensions[column].width = width

        # 寫入標題
        header_font = Font(size=18, bold=True)
        for col_index, col_value in enumerate(header, start=1):
            sheet.cell(row=2, column=col_index, value=col_value).font = header_font
            sheet.cell(row=2, column=col_index).alignment = Alignment(horizontal='center')

        # 設定表格邊框
        border_style = Side(border_style="thin", color="000000")
        border = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)

        # 設定標題和標頭行的邊框
        for cell in sheet[1]:
            cell.border = border

        for cell in sheet[2]:
            cell.border = border


        # 調整行高
        row_heights = {
            1: 35,
            2: 25
        }
        for row, height in row_heights.items():
            sheet.row_dimensions[row].height = height

        # 寫入資料
        for row in month_data:
            date = row['西元日期']
            day = int(date[6:8])
            weekday = row['星期']
            holiday = int(row['是否放假'])
            remark = row['備註']

            # 判斷是否放假，並將整行文字改為紅色
            if holiday == 2:
                font = Font(color='FF0000', size=15, bold=True)
                remark = remark.upper()
            else:
                font = Font(size=15)

            # 修改日期格式
            date = f'{month}月{day}日'

            # 判斷備註是否為「補行上班」，並將整行文字改為藍色
            if remark == '補行上班':
                font.color = colors.BLUE

            # 移除 remark
            row.pop('備註', None)

            # 寫入資料到Excel
            data_row = [date, weekday]
            for col in header[2:]:  # 略過日期和星期
                data_row.append(None)
            sheet.append(data_row)
            
            # 識別 remark 應該插入的位置
            remark_index = header.index('備註')
            remark_col_letter = chr(ord('A') + remark_index)

            sheet[f'{remark_col_letter}{sheet.max_row}'] = remark

            # 設定每個儲存格的邊框
            for i, cell in enumerate(sheet[sheet.max_row]):
                cell.font = font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

        # 調整 title 合併的範圍
        title_col_letter_start = 'A'
        title_col_letter_end = chr(ord('A') + len(header) - 1)
        title_range = f'{title_col_letter_start}1:{title_col_letter_end}1'
        sheet.merge_cells(title_range)

        # 寫入 title
        sheet[title_col_letter_start + '1'] = title
        sheet[title_col_letter_start + '1'].font = Font(size=30, bold=True)
        sheet[title_col_letter_start + '1'].alignment = Alignment(horizontal='center')

    # 刪除預設的工作表
    workbook.remove(workbook['Sheet'])

    # 儲存Excel檔案
    workbook.save(out_path)

# 使用函數生成 Excel 檔案
# csv_path, out_path, title, header, column_widths
generate_excel('assets/113年中華民國政府行政機關辦公日曆表.csv', 'out/', '113_TCSN_傳輸機房溫度檢查表', ['日期', '星期', '量測點E12溫度', '量測點J09溫度', '量測者', '備註'], {'A': 15, 'B': 10, 'C': 25, 'D': 25, 'E': 10, 'F': 25})
generate_excel('assets/113年中華民國政府行政機關辦公日曆表.csv', 'out/', '113_TCHL_傳輸機房溫度檢查表', ['日期', '星期', 'POSS量測點4F', 'POSS量測點2F','量測者', '備註'], {'A': 15, 'B': 10, 'C': 25, 'D': 25, 'E': 10, 'F': 25})
generate_excel('assets/113年中華民國政府行政機關辦公日曆表.csv', 'out/', '113_TCFC_傳輸機房溫度檢查表', ['日期', '星期', 'POSS量測點5F', '量測者', '備註'], {'A': 15, 'B': 10, 'C': 25, 'D': 10, 'E': 25})
